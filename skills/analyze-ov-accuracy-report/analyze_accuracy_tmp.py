from openpyxl import load_workbook
from collections import Counter, defaultdict
from pathlib import Path

path = Path(r"C:\projects\sourcecode\test\2026-02-24_08-18_WW08_N_G_NG_weekly.xlsx")
wb = load_workbook(path, data_only=True)
ws = wb[wb.sheetnames[0]]

rows = list(ws.iter_rows(values_only=True))
if not rows:
    print("No data")
    raise SystemExit(0)

header = [str(c).strip() if c is not None else "" for c in rows[0]]
lower_header = [h.lower() for h in header]

col_index = {name.lower(): i for i, name in enumerate(header)}

def get_col(name):
    return col_index.get(name.lower())

topology_idx = get_col("topology")
precision_idx = get_col("precision")
device_idx = get_col("device")

metric_indices = [i for i, name in enumerate(header) if "metric" in name.lower()]
metric_status_indices = [i for i, name in enumerate(header) if "metric" in name.lower() and "status" in name.lower()]

def find_status_idx(tag):
    for i, name in enumerate(lower_header):
        if "metric" in name and "status" in name and tag in name:
            return i
    return None

ww08_status_idx = find_status_idx("ww08")
ww07_status_idx = find_status_idx("ww07")
rc3_status_idx = find_status_idx("2026.0.0-rc3")

records = []
records_by_device = defaultdict(list)
ww08_downgrade_rc3_passed = 0
ww08_rc3_records = []
ww08_rc3_records_by_device = defaultdict(list)
ww08_rc3_rows_by_device = defaultdict(list)
ww08_new_downgrades = 0
ww08_new_records = []
ww08_improvements = 0
ww08_improve_records = []

for r in rows[1:]:
    if all(v is None for v in r):
        continue
    device_val = str(r[device_idx]) if device_idx is not None else ""
    device_lower = device_val.lower()
    if "cpu" in device_lower:
        device_group = "CPU"
    elif "gpu" in device_lower:
        device_group = "GPU"
    else:
        device_group = "OTHER"

    topo = r[topology_idx] if topology_idx is not None else None
    prec = r[precision_idx] if precision_idx is not None else None

    fail_metrics = []
    for mi in metric_indices:
        val = r[mi]
        if val is None:
            continue
        if "downgrade" in str(val).lower():
            fail_metrics.append(header[mi])

    if ww08_status_idx is not None and rc3_status_idx is not None:
        ww08_val = r[ww08_status_idx]
        rc3_val = r[rc3_status_idx]
        if ww08_val is not None and rc3_val is not None:
            if "downgrade" in str(ww08_val).lower() and "passed" in str(rc3_val).lower():
                ww08_downgrade_rc3_passed += 1
                ww08_rc3_records.append((topo, prec))
                ww08_rc3_records_by_device[device_group].append((topo, prec))
                ww08_rc3_rows_by_device[device_group].append(
                    (topo, device_val, prec, ww08_val, rc3_val)
                )

    if ww08_status_idx is not None and ww07_status_idx is not None:
        ww08_val = r[ww08_status_idx]
        ww07_val = r[ww07_status_idx]
        if ww08_val is not None and ww07_val is not None:
            ww08_is_downgrade = "downgrade" in str(ww08_val).lower()
            ww07_is_passed = "passed" in str(ww07_val).lower()
            ww07_is_downgrade = "downgrade" in str(ww07_val).lower()
            ww08_is_passed = "passed" in str(ww08_val).lower()
            if ww08_is_downgrade and ww07_is_passed:
                ww08_new_downgrades += 1
                ww08_new_records.append((topo, prec))
            if ww08_is_passed and ww07_is_downgrade:
                ww08_improvements += 1
                ww08_improve_records.append((topo, prec))

    if not fail_metrics:
        continue

    if device_group == "CPU":
        for m in fail_metrics:
            records.append((topo, prec, m))

    for m in fail_metrics:
        records_by_device[device_group].append((topo, prec, m))

cnt_topo = Counter([r[0] for r in records])
cnt_prec = Counter([r[1] for r in records])
cnt_metric = Counter([r[2] for r in records])

print("cpu_failures", len(records))
print("cpu_topo", cnt_topo.most_common(10))
print("cpu_prec", cnt_prec.most_common(10))
print("cpu_metric", cnt_metric.most_common(10))
print("ww08 downgrade + rc3 passed", ww08_downgrade_rc3_passed)
print("ww08 new downgrades vs ww07", ww08_new_downgrades)
print("ww08 improvements vs ww07", ww08_improvements)

if ww08_rc3_records:
    combo_ww08 = Counter(ww08_rc3_records)
    print("ww08 downgrade + rc3 passed by topology+precision (top 20)")
    for (topo, prec), count in combo_ww08.most_common(20):
        print(f"{count}\t{topo}\t{prec}")
    print("ww08 downgrade + rc3 passed by device")
    for device_group in ("CPU", "GPU", "OTHER"):
        total = len(ww08_rc3_records_by_device.get(device_group, []))
        print(f"{device_group}\t{total}")
    print("ww08 downgrade + rc3 passed by device and topology (top 10 per device)")
    for device_group in ("CPU", "GPU", "OTHER"):
        device_records = ww08_rc3_records_by_device.get(device_group, [])
        if not device_records:
            continue
        print(f"{device_group} top 10")
        combo_device = Counter(device_records)
        for (topo, prec), count in combo_device.most_common(10):
            print(f"{count}\t{topo}\t{prec}")
    cpu_rows = ww08_rc3_rows_by_device.get("CPU", [])
    if cpu_rows:
        print("ww08 downgrade + rc3 passed CPU rows")
        print("topology\tdevice\tprecision\tww08_status\trc3_status")
        for topo, device_val, prec, ww08_val, rc3_val in cpu_rows:
            print(f"{topo}\t{device_val}\t{prec}\t{ww08_val}\t{rc3_val}")

if ww08_new_records:
    combo_ww08_new = Counter(ww08_new_records)
    print("ww08 new downgrades vs ww07 by topology+precision (top 20)")
    for (topo, prec), count in combo_ww08_new.most_common(20):
        print(f"{count}\t{topo}\t{prec}")

if ww08_improve_records:
    combo_ww08_improve = Counter(ww08_improve_records)
    print("ww08 improvements vs ww07 by topology+precision (top 20)")
    for (topo, prec), count in combo_ww08_improve.most_common(20):
        print(f"{count}\t{topo}\t{prec}")

combo = Counter(records)
print("cpu_top_50_topo_prec_metric")
for (topo, prec, metric), count in combo.most_common(50):
    print(f"{count}\t{topo}\t{prec}\t{metric}")

if records:
    top_topos = [t for t, _ in cnt_topo.most_common(10)]
    if top_topos:
        print("topology_cross_device_counts")
        for topo in top_topos:
            cpu_count = sum(1 for r in records_by_device["CPU"] if r[0] == topo)
            gpu_count = sum(1 for r in records_by_device["GPU"] if r[0] == topo)
            other_count = sum(1 for r in records_by_device["OTHER"] if r[0] == topo)
            print(f"{topo}\tCPU:{cpu_count}\tGPU:{gpu_count}\tOTHER:{other_count}")
